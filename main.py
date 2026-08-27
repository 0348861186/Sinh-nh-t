import io
import re
import json
import hashlib
import urllib.request
from pathlib import Path
from collections import defaultdict

import streamlit as st
import openpyxl
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from PIL import Image, ImageEnhance, ImageFilter, ImageOps

# ============================================================
# LOCAL TRANSLATION + OCR
# Không dùng AI/API/quota
# ============================================================
import argostranslate.package
import argostranslate.translate

try:
    import easyocr
except Exception:
    easyocr = None

try:
    import fitz  # PyMuPDF
except Exception:
    fitz = None


# ============================================================
# CẤU HÌNH STREAMLIT
# ============================================================
st.set_page_config(
    page_title="Hệ Thống Dịch Bảng Chấm Công Thông Minh",
    page_icon="🌐",
    layout="wide",
)

st.title("🌐 Dịch & Xuất Bảng Chấm Công Song Ngữ")
st.caption(
    "Local 100% | Không dùng AI/API | OCR đa tầng | Từ điển chuyên ngành | "
    "Translation Memory | Tự nhận diện cột"
)


# ============================================================
# THƯ MỤC LOCAL
# ============================================================
BASE_DIR = Path(".")
MODEL_DIR = BASE_DIR / ".argos_models"
MEMORY_FILE = BASE_DIR / "translation_memory.json"
MODEL_DIR.mkdir(exist_ok=True)


# ============================================================
# TỪ ĐIỂN CHUYÊN NGÀNH
# Có thể bổ sung tùy ý.
# Cụm dài được ưu tiên trước cụm ngắn.
# ============================================================
GLOSSARY_ZH_VI = {
    # Hệ thống chung
    "正式工": "Công nhân chính thức",
    "临时工": "Công nhân thời vụ",
    "部门": "Bộ phận",
    "车间": "Xưởng",
    "班组": "Tổ sản xuất",
    "生产线": "Chuyền sản xuất",
    "开几台机": "Số máy đang chạy",
    "开机": "Chạy máy",
    "机台": "Máy",
    "机器": "Máy",
    "备注": "Ghi chú",
    "总计": "Tổng cộng",
    "合计": "Tổng cộng",
    "姓名": "Họ tên",
    "班次": "Ca làm việc",
    "早班": "Ca sáng",
    "白班": "Ca ngày",
    "晚班": "Ca đêm",
    "夜班": "Ca đêm",
    "休息": "Nghỉ",
    "出勤": "Đi làm",
    "缺勤": "Vắng mặt",
    "迟到": "Đi muộn",
    "早退": "Về sớm",
    "加班": "Tăng ca",
    "请假": "Xin nghỉ",
    "病假": "Nghỉ bệnh",
    "事假": "Nghỉ việc riêng",

    # Sản xuất / nhà máy
    "裁断": "Cắt",
    "裁断组": "Tổ cắt",
    "针车": "May",
    "针车组": "Tổ may",
    "成型": "Thành hình",
    "成型组": "Tổ thành hình",
    "包装": "Đóng gói",
    "包装组": "Tổ đóng gói",
    "品检": "Kiểm hàng",
    "品质": "Chất lượng",
    "仓库": "Kho",
    "仓管": "Quản lý kho",
    "维修": "Bảo trì",
    "机修": "Bảo trì máy",
    "行政": "Hành chính",
    "人事": "Nhân sự",
    "财务": "Tài vụ",
    "采购": "Mua hàng",
    "生管": "Quản lý sản xuất",
    "生产": "Sản xuất",
}


GLOSSARY_VI_ZH = {v: k for k, v in GLOSSARY_ZH_VI.items()}


# ============================================================
# CỤM TỪ OCR THƯỜNG NHẦM
# Có thể mở rộng theo file thực tế.
# ============================================================
OCR_NORMALIZATION_ZH = {
    "正式エ": "正式工",
    "正式T": "正式工",
    "正式丁": "正式工",
    "临时エ": "临时工",
    "臨時工": "临时工",
    "部門": "部门",
    "備注": "备注",
    "總計": "总计",
    "合計": "合计",
    "開機": "开机",
    "機台": "机台",
    "車間": "车间",
    "針車": "针车",
    "裁斷": "裁断",
    "成型組": "成型组",
    "裁斷組": "裁断组",
}


# ============================================================
# ARGOS MODELS
# Ưu tiên direct zh-vi / vi-zh nếu có.
# Fallback qua English nếu direct chưa có.
# ============================================================
ARGOS_PACKAGES = {
    "zh_en": "translate-zh_en-1_9.argosmodel",
    "en_zh": "translate-en_zh-1_9.argosmodel",
    "vi_en": "translate-vi_en-1_9.argosmodel",
    "en_vi": "translate-en_vi-1_9.argosmodel",
}

ARGOS_BASE_URL = "https://data.argosopentech.com/argospm/v1/"


# ============================================================
# CACHE / SESSION
# ============================================================
if "translation_stats" not in st.session_state:
    st.session_state.translation_stats = {
        "glossary": 0,
        "memory": 0,
        "argos": 0,
        "unchanged": 0,
    }


# ============================================================
# TEXT UTILITIES
# ============================================================
def normalize_text(text):
    if text is None:
        return ""

    text = str(text)
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    text = text.replace("\u3000", " ")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def normalize_ocr_text(text):
    text = normalize_text(text)
    if not text:
        return ""

    for src, dst in sorted(
        OCR_NORMALIZATION_ZH.items(),
        key=lambda x: len(x[0]),
        reverse=True,
    ):
        text = text.replace(src, dst)

    # Một số lỗi OCR khoảng trắng trong cụm Trung
    text = re.sub(r"(?<=正)\s+(?=式)", "", text)
    text = re.sub(r"(?<=临)\s+(?=时)", "", text)
    text = re.sub(r"(?<=开)\s+(?=机)", "", text)
    text = re.sub(r"(?<=部)\s+(?=门)", "", text)

    return text.strip()


def is_number_like(text):
    text = normalize_text(text)
    if not text:
        return False

    if re.fullmatch(r"[-+]?\d+(?:[.,]\d+)?", text):
        return True

    if re.fullmatch(r"[\d\s.,:/\\\-+%]+", text):
        return True

    if re.fullmatch(r"[\d\s.,:/\\\-+%]+(?:人|台|个|名|小时|天)", text):
        return True

    return False


def should_translate(text):
    text = normalize_text(text)
    if not text:
        return False

    if is_number_like(text):
        return False

    # Mã dạng ABC-123 / 123-456 không nên đưa vào MT
    if re.fullmatch(r"[A-Za-z0-9_\-/.:]+", text):
        if any(c.isalpha() for c in text) and any(c.isdigit() for c in text):
            return False

    return True


# ============================================================
# GLOSSARY
# ============================================================
def apply_glossary_exact(text, mode):
    text = normalize_text(text)
    if not text:
        return None

    glossary = GLOSSARY_ZH_VI if mode == "Trung ➔ Việt" else GLOSSARY_VI_ZH
    return glossary.get(text)


def apply_glossary_replace(text, mode):
    text = normalize_text(text)
    if not text:
        return text

    glossary = GLOSSARY_ZH_VI if mode == "Trung ➔ Việt" else GLOSSARY_VI_ZH

    # Cụm dài trước để tránh thay một phần sai.
    result = text
    for src in sorted(glossary.keys(), key=len, reverse=True):
        result = result.replace(src, glossary[src])

    return result


def glossary_for_context(context, mode):
    if mode == "Trung ➔ Việt":
        base = GLOSSARY_ZH_VI
    else:
        base = GLOSSARY_VI_ZH

    # Có thể mở rộng riêng theo context về sau.
    return base


def apply_context_glossary(text, mode, context=None):
    text = normalize_text(text)
    if not text:
        return text

    glossary = glossary_for_context(context, mode)

    for src in sorted(glossary.keys(), key=len, reverse=True):
        text = text.replace(src, glossary[src])

    return text


# ============================================================
# TRANSLATION MEMORY
# ============================================================
def load_translation_memory():
    if not MEMORY_FILE.exists():
        return {
            "zh_vi": {},
            "vi_zh": {},
        }

    try:
        data = json.loads(MEMORY_FILE.read_text(encoding="utf-8"))
        if not isinstance(data, dict):
            raise ValueError("Translation memory không hợp lệ.")

        data.setdefault("zh_vi", {})
        data.setdefault("vi_zh", {})
        return data
    except Exception:
        return {
            "zh_vi": {},
            "vi_zh": {},
        }


def save_translation_memory(memory):
    tmp = MEMORY_FILE.with_suffix(".tmp")

    tmp.write_text(
        json.dumps(memory, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    tmp.replace(MEMORY_FILE)


def memory_key(mode):
    return "zh_vi" if mode == "Trung ➔ Việt" else "vi_zh"


def memory_lookup(text, mode):
    text = normalize_text(text)
    if not text:
        return None

    memory = load_translation_memory()
    return memory.get(memory_key(mode), {}).get(text)


def memory_save(source, target, mode):
    source = normalize_text(source)
    target = normalize_text(target)

    if not source or not target:
        return

    if source == target:
        return

    memory = load_translation_memory()
    key = memory_key(mode)

    if source not in memory[key]:
        memory[key][source] = target
        try:
            save_translation_memory(memory)
        except Exception:
            pass


# ============================================================
# ARGOS
# ============================================================
@st.cache_resource
def get_argos_languages():
    try:
        return argostranslate.translate.get_installed_languages()
    except Exception:
        return []


def clear_argos_language_cache():
    try:
        get_argos_languages.clear()
    except Exception:
        pass


def find_argos_language(code):
    for lang in get_argos_languages():
        if lang.code == code:
            return lang
    return None


def find_argos_translation(from_code, to_code):
    installed_languages = get_argos_languages()

    from_lang = None
    to_lang = None

    for lang in installed_languages:
        if lang.code == from_code:
            from_lang = lang
        if lang.code == to_code:
            to_lang = lang

    if from_lang is None or to_lang is None:
        return None

    try:
        return from_lang.get_translation(to_lang)
    except Exception:
        return None


def download_argos_package(package_name):
    MODEL_DIR.mkdir(exist_ok=True)

    target = MODEL_DIR / package_name

    if target.exists() and target.stat().st_size > 0:
        return target

    url = ARGOS_BASE_URL + package_name

    try:
        urllib.request.urlretrieve(url, str(target))
        return target
    except Exception as e:
        if target.exists():
            try:
                target.unlink()
            except Exception:
                pass

        raise RuntimeError(f"Không thể tải model Argos: {e}")


def install_argos_package(from_code, to_code):
    package_key = f"{from_code}_{to_code}"

    if package_key not in ARGOS_PACKAGES:
        raise RuntimeError(
            f"Chưa cấu hình model Argos {from_code} -> {to_code}"
        )

    package_path = download_argos_package(ARGOS_PACKAGES[package_key])

    try:
        argostranslate.package.install_from_path(str(package_path))
    except Exception as e:
        message = str(e).lower()

        if "already installed" not in message:
            raise

    clear_argos_language_cache()


@st.cache_resource
def initialize_translation_models():
    # Chỉ tải những model cần thiết cho fallback.
    required_pairs = [
        ("zh", "en"),
        ("en", "zh"),
        ("vi", "en"),
        ("en", "vi"),
    ]

    for from_code, to_code in required_pairs:
        if find_argos_translation(from_code, to_code) is None:
            install_argos_package(from_code, to_code)


def translate_direct(text, from_code, to_code):
    text = normalize_text(text)

    if not text or from_code == to_code:
        return text

    translation = find_argos_translation(from_code, to_code)

    if translation is None:
        install_argos_package(from_code, to_code)
        translation = find_argos_translation(from_code, to_code)

    if translation is None:
        return text

    try:
        result = translation.translate(text)
        result = normalize_text(result)
        return result if result else text
    except Exception:
        return text


# ============================================================
# POST PROCESS TRANSLATION
# ============================================================
def postprocess_translation(source, translated, mode):
    source = normalize_text(source)
    translated = normalize_text(translated)

    if not translated:
        return source

    # Không để model phá số đơn giản.
    if is_number_like(source):
        return source

    # Sau khi model dịch, ưu tiên sửa các thuật ngữ rõ ràng.
    if mode == "Trung ➔ Việt":
        glossary = GLOSSARY_ZH_VI
    else:
        glossary = GLOSSARY_VI_ZH

    # Chỉ thay trong output khi output còn chứa source phrase.
    for src, dst in sorted(glossary.items(), key=lambda x: len(x[0]), reverse=True):
        if src in translated:
            translated = translated.replace(src, dst)

    return translated.strip()


# ============================================================
# SMART TRANSLATOR
# ============================================================
def translate_text(text, mode, context=None):
    text = normalize_ocr_text(text)

    if not text:
        return ""

    if not should_translate(text):
        st.session_state.translation_stats["unchanged"] += 1
        return text

    # --------------------------------------------------------
    # 1. Exact glossary
    # --------------------------------------------------------
    exact = apply_glossary_exact(text, mode)

    if exact is not None:
        st.session_state.translation_stats["glossary"] += 1
        return exact

    # --------------------------------------------------------
    # 2. Translation Memory
    # --------------------------------------------------------
    memory_result = memory_lookup(text, mode)

    if memory_result:
        st.session_state.translation_stats["memory"] += 1
        return memory_result

    # --------------------------------------------------------
    # 3. Context glossary
    # --------------------------------------------------------
    context_text = apply_context_glossary(text, mode, context)

    # Nếu sau glossary toàn bộ đã thành bản dịch thì dùng luôn.
    if context_text != text:
        # Nếu còn ký tự nguồn thì vẫn cho MT xử lý phần còn lại.
        if mode == "Trung ➔ Việt":
            has_zh = bool(re.search(r"[\u3400-\u9fff]", context_text))
        else:
            has_zh = False

        if not has_zh:
            result = normalize_text(context_text)
            memory_save(text, result, mode)
            st.session_state.translation_stats["glossary"] += 1
            return result

    # --------------------------------------------------------
    # 4. Direct model nếu có
    # --------------------------------------------------------
    if mode == "Trung ➔ Việt":
        direct = find_argos_translation("zh", "vi")

        if direct is not None:
            try:
                result = normalize_text(direct.translate(context_text))
                if result:
                    result = postprocess_translation(text, result, mode)
                    memory_save(text, result, mode)
                    st.session_state.translation_stats["argos"] += 1
                    return result
            except Exception:
                pass

        # ----------------------------------------------------
        # 5. Fallback zh -> en -> vi
        # ----------------------------------------------------
        english = translate_direct(context_text, "zh", "en")

        # Nếu zh -> en không tạo được kết quả hữu ích
        if not english or english == context_text:
            return text

        vietnamese = translate_direct(english, "en", "vi")
        result = normalize_text(vietnamese)

    else:
        direct = find_argos_translation("vi", "zh")

        if direct is not None:
            try:
                result = normalize_text(direct.translate(context_text))
                if result:
                    result = postprocess_translation(text, result, mode)
                    memory_save(text, result, mode)
                    st.session_state.translation_stats["argos"] += 1
                    return result
            except Exception:
                pass

        # ----------------------------------------------------
        # 5. Fallback vi -> en -> zh
        # ----------------------------------------------------
        english = translate_direct(context_text, "vi", "en")

        if not english or english == context_text:
            return text

        chinese = translate_direct(english, "en", "zh")
        result = normalize_text(chinese)

    result = postprocess_translation(text, result, mode)

    if result:
        memory_save(text, result, mode)
        st.session_state.translation_stats["argos"] += 1
        return result

    st.session_state.translation_stats["unchanged"] += 1
    return text


# ============================================================
# OCR PREPROCESSING
# ============================================================
def resize_for_ocr(image):
    if image.mode != "RGB":
        image = image.convert("RGB")

    width, height = image.size

    # Không phóng quá lớn gây nặng RAM.
    if width < 1800:
        scale = 2.0
    elif width < 3000:
        scale = 1.5
    else:
        scale = 1.0

    if scale != 1.0:
        image = image.resize(
            (int(width * scale), int(height * scale)),
            Image.Resampling.LANCZOS,
        )

    return image


def create_ocr_variants(image):
    base = resize_for_ocr(image)

    variants = []

    # 1. Original
    variants.append(("original", base))

    # 2. Contrast + sharp
    enhanced = ImageEnhance.Contrast(base).enhance(1.45)
    enhanced = ImageEnhance.Sharpness(enhanced).enhance(1.6)
    variants.append(("enhanced", enhanced))

    # 3. Grayscale + autocontrast
    gray = ImageOps.grayscale(base)
    gray = ImageOps.autocontrast(gray)
    variants.append(("gray", gray))

    # 4. Sharpen grayscale
    sharp = gray.filter(ImageFilter.SHARPEN)
    sharp = ImageEnhance.Contrast(sharp).enhance(1.35)
    variants.append(("sharp_gray", sharp))

    # 5. Mild threshold
    arr = np.array(gray)
    threshold = 175
    binary = np.where(arr > threshold, 255, 0).astype(np.uint8)
    binary_img = Image.fromarray(binary)
    variants.append(("threshold", binary_img))

    return variants


@st.cache_resource
def get_ocr():
    if easyocr is None:
        raise RuntimeError(
            "Chưa cài EasyOCR. Hãy cài easyocr và torch trong requirements.txt."
        )

    # Trung giản thể + English
    return easyocr.Reader(
        ["ch_sim", "en"],
        gpu=False,
        verbose=False,
    )


# ============================================================
# OCR GEOMETRY
# ============================================================
def get_box_geometry(box):
    if not box:
        return None

    try:
        xs = [float(p[0]) for p in box]
        ys = [float(p[1]) for p in box]

        return (
            min(xs),
            min(ys),
            max(xs),
            max(ys),
        )
    except Exception:
        return None


def box_iou(box_a, box_b):
    if not box_a or not box_b:
        return 0.0

    ax1, ay1, ax2, ay2 = box_a
    bx1, by1, bx2, by2 = box_b

    ix1 = max(ax1, bx1)
    iy1 = max(ay1, by1)
    ix2 = min(ax2, bx2)
    iy2 = min(ay2, by2)

    iw = max(0.0, ix2 - ix1)
    ih = max(0.0, iy2 - iy1)

    inter = iw * ih

    area_a = max(0.0, ax2 - ax1) * max(0.0, ay2 - ay1)
    area_b = max(0.0, bx2 - bx1) * max(0.0, by2 - by1)

    union = area_a + area_b - inter

    if union <= 0:
        return 0.0

    return inter / union


def deduplicate_ocr_results(items):
    prepared = []

    for item in items:
        geo = get_box_geometry(item.get("box"))

        if not geo:
            continue

        text = normalize_ocr_text(item.get("text", ""))

        if not text:
            continue

        prepared.append(
            {
                **item,
                "text": text,
                "geo": geo,
                "score": float(item.get("score", 0.0)),
            }
        )

    # Confidence cao trước
    prepared.sort(key=lambda x: x["score"], reverse=True)

    kept = []

    for item in prepared:
        duplicate = False

        for old in kept:
            iou = box_iou(item["geo"], old["geo"])

            if iou >= 0.45:
                duplicate = True
                break

            # Hai box gần như cùng vị trí
            ax1, ay1, ax2, ay2 = item["geo"]
            bx1, by1, bx2, by2 = old["geo"]

            acx = (ax1 + ax2) / 2
            acy = (ay1 + ay2) / 2

            bcx = (bx1 + bx2) / 2
            bcy = (by1 + by2) / 2

            aw = max(ax2 - ax1, 1)
            ah = max(ay2 - ay1, 1)

            if (
                abs(acx - bcx) <= aw * 0.35
                and abs(acy - bcy) <= ah * 0.5
                and item["text"] == old["text"]
            ):
                duplicate = True
                break

        if not duplicate:
            kept.append(item)

    kept.sort(
        key=lambda x: (
            (x["geo"][1] + x["geo"][3]) / 2,
            x["geo"][0],
        )
    )

    for item in kept:
        item.pop("geo", None)

    return kept


def ocr_image(image):
    reader = get_ocr()

    variants = create_ocr_variants(image)

    all_results = []

    for variant_name, variant in variants:
        try:
            results = reader.readtext(
                np.array(variant),
                detail=1,
                paragraph=False,
                text_threshold=0.45,
                low_text=0.25,
                link_threshold=0.25,
                mag_ratio=1.0,
            )
        except Exception:
            continue

        for bbox, text, prob in results:
            text = normalize_ocr_text(text)

            if not text:
                continue

            # Ngưỡng mềm để không bỏ mất chữ nhỏ.
            if float(prob) < 0.25:
                continue

            all_results.append(
                {
                    "text": text,
                    "score": float(prob),
                    "box": bbox,
                    "variant": variant_name,
                }
            )

    return deduplicate_ocr_results(all_results)


# ============================================================
# OCR LINE GROUPING
# ============================================================
def group_ocr_lines(items):
    prepared = []

    for item in items:
        geo = get_box_geometry(item.get("box"))

        if not geo:
            continue

        x1, y1, x2, y2 = geo

        prepared.append(
            {
                **item,
                "x1": x1,
                "y1": y1,
                "x2": x2,
                "y2": y2,
                "cx": (x1 + x2) / 2,
                "cy": (y1 + y2) / 2,
                "height": max(y2 - y1, 1),
            }
        )

    prepared.sort(key=lambda x: (x["cy"], x["x1"]))

    lines = []

    for item in prepared:
        placed = False

        for line in lines:
            avg_y = sum(x["cy"] for x in line) / len(line)
            avg_h = sum(x["height"] for x in line) / len(line)

            # Ngưỡng động theo chiều cao chữ.
            tolerance = max(10, avg_h * 0.65)

            if abs(item["cy"] - avg_y) <= tolerance:
                line.append(item)
                placed = True
                break

        if not placed:
            lines.append([item])

    for line in lines:
        line.sort(key=lambda x: x["x1"])

    return lines


def line_to_text(line):
    return " ".join(
        normalize_ocr_text(item.get("text", ""))
        for item in line
        if normalize_ocr_text(item.get("text", ""))
    )


# ============================================================
# DATE EXTRACTION
# ============================================================
def extract_date(text):
    text = normalize_text(text)

    patterns = [
        r"(\d{4})[-/.年](\d{1,2})[-/.月](\d{1,2})日?",
        r"(\d{1,2})[-/.](\d{1,2})[-/.](\d{4})",
        r"(\d{4})年(\d{1,2})月(\d{1,2})日",
    ]

    for pattern in patterns:
        for match in re.finditer(pattern, text):
            groups = match.groups()

            try:
                if len(groups[0]) == 4:
                    year = int(groups[0])
                    month = int(groups[1])
                    day = int(groups[2])
                else:
                    day = int(groups[0])
                    month = int(groups[1])
                    year = int(groups[2])

                if 1 <= month <= 12 and 1 <= day <= 31:
                    return f"{year:04d}-{month:02d}-{day:02d}"
            except Exception:
                pass

    return ""


# ============================================================
# NUMBER CLEANING
# ============================================================
def clean_number(text):
    if text is None:
        return ""

    text = str(text).strip()
    text = text.replace("，", ".").replace(",", ".")

    # Bỏ ký tự OCR phổ biến xung quanh số
    text = text.replace("O", "0").replace("o", "0")

    match = re.search(r"-?\d+(?:\.\d+)?", text)

    if not match:
        return text

    try:
        val = float(match.group(0))

        if val.is_integer():
            return int(val)

        return val
    except Exception:
        return text


# ============================================================
# HEADER DETECTION
# ============================================================
HEADER_KEYWORDS = [
    "部门",
    "开几台",
    "开机",
    "机台",
    "正式",
    "临时",
    "备注",
    "姓名",
    "班次",
    "bộ phận",
    "số máy",
    "máy",
    "chính thức",
    "thời vụ",
    "ghi chú",
    "stt",
    "tên",
    "họ tên",
]


def detect_header_line(lines):
    best_idx = None
    best_score = 0.0

    for idx, line in enumerate(lines):
        text = line_to_text(line).lower()

        if not text:
            continue

        score = 0.0

        for kw in HEADER_KEYWORDS:
            if kw.lower() in text:
                # Từ khóa dài đáng tin hơn.
                score += max(1.0, len(kw) / 4)

        # Header thường có nhiều box/cột.
        if len(line) >= 3:
            score += 1.5

        if score > best_score:
            best_score = score
            best_idx = idx

    return best_idx


# ============================================================
# COLUMN CLASSIFICATION
# ============================================================
def classify_column_dynamic(col_name_lower):
    name = normalize_text(col_name_lower).lower()

    # Thứ tự rất quan trọng: từ cụ thể -> từ chung.
    if any(
        k in name
        for k in [
            "备注",
            "ghi chú",
            "chú thích",
            "remark",
            "note",
        ]
    ):
        return "remark"

    if any(
        k in name
        for k in [
            "正式工",
            "正式",
            "chính thức",
            "cố định",
            "formal",
        ]
    ):
        return "formal"

    if any(
        k in name
        for k in [
            "临时工",
            "临时",
            "thời vụ",
            "lao động phụ",
            "temp",
        ]
    ):
        return "temp"

    if any(
        k in name
        for k in [
            "开几台机",
            "开机",
            "机台",
            "机器",
            "số máy",
            "số máy mở",
            "máy",
            "machine",
        ]
    ):
        return "machines"

    if any(
        k in name
        for k in [
            "部门",
            "bộ phận",
            "phòng ban",
            "phòng",
            "xưởng",
            "车间",
            "department",
        ]
    ):
        return "dept_src"

    # Không dùng "tên" đơn độc làm dept vì dễ nhầm "tên máy".
    if any(
        k in name
        for k in [
            "tên bộ phận",
            "tên phòng",
            "tên xưởng",
            "department name",
        ]
    ):
        return "dept_src"

    return "unknown"


def infer_column_boundaries(columns):
    """
    columns: [{name, x}]
    Tạo boundary giữa tâm các cột.
    """
    if not columns:
        return []

    columns = sorted(columns, key=lambda x: x["x"])

    boundaries = []

    for i, col in enumerate(columns):
        if i == 0:
            left = float("-inf")
        else:
            left = (columns[i - 1]["x"] + col["x"]) / 2

        if i == len(columns) - 1:
            right = float("inf")
        else:
            right = (col["x"] + columns[i + 1]["x"]) / 2

        boundaries.append(
            {
                **col,
                "left": left,
                "right": right,
            }
        )

    return boundaries


def find_column_by_x(boundaries, x):
    if not boundaries:
        return None

    for col in boundaries:
        if col["left"] <= x < col["right"]:
            return col

    return min(
        boundaries,
        key=lambda c: abs(c["x"] - x),
    )


# ============================================================
# STT
# ============================================================
def extract_stt(text):
    text = normalize_text(text)

    patterns = [
        r"^\s*(\d+)[.)]?\s*$",
        r"^\s*(\d+)",
    ]

    for pattern in patterns:
        match = re.match(pattern, text)

        if match:
            try:
                return int(match.group(1))
            except Exception:
                pass

    return None


# ============================================================
# PARSE ATTENDANCE ROWS
# ============================================================
def parse_attendance_rows(lines, header_index):
    if not lines:
        return []

    if header_index is None:
        header_index = 0

    if header_index >= len(lines):
        header_index = 0

    header_line = lines[header_index]

    columns = []

    for item in header_line:
        name = normalize_ocr_text(item.get("text", "")).lower()

        if not name:
            continue

        columns.append(
            {
                "name": name,
                "x": item["cx"],
                "type": classify_column_dynamic(name),
            }
        )

    columns.sort(key=lambda x: x["x"])

    boundaries = infer_column_boundaries(columns)

    rows = []

    for line in lines[header_index + 1:]:
        if not line:
            continue

        line_text = line_to_text(line)

        if not line_text:
            continue

        lower_line = line_text.lower()

        # Bỏ dòng tổng.
        if any(
            kw in lower_line
            for kw in [
                "一共",
                "总计",
                "合计",
                "tổng cộng",
                "total",
            ]
        ):
            continue

        # Bỏ dòng tiêu đề lặp lại ở trang sau.
        if detect_header_line([line]) is not None:
            header_score = sum(
                1
                for kw in HEADER_KEYWORDS
                if kw.lower() in lower_line
            )

            if header_score >= 2:
                continue

        # STT
        stt = None

        if line:
            stt = extract_stt(line[0].get("text", ""))

        if stt is None:
            stt = extract_stt(line_text)

        if stt is None:
            stt = len(rows) + 1

        row = {
            "stt": stt,
            "dept_src": "",
            "dept_tgt": "",
            "machines": "",
            "formal": "",
            "temp": "",
            "remark": "",
        }

        unknown_items = []

        for item in line:
            text = normalize_ocr_text(item.get("text", ""))

            if not text:
                continue

            col = find_column_by_x(boundaries, item["cx"])

            if col is None:
                unknown_items.append(text)
                continue

            col_type = col["type"]

            if col_type == "dept_src":
                row["dept_src"] = (
                    row["dept_src"] + " " + text
                ).strip()

            elif col_type == "machines":
                cleaned = clean_number(text)

                if row["machines"] == "":
                    row["machines"] = cleaned

            elif col_type == "formal":
                cleaned = clean_number(text)

                if row["formal"] == "":
                    row["formal"] = cleaned

            elif col_type == "temp":
                cleaned = clean_number(text)

                if row["temp"] == "":
                    row["temp"] = cleaned

            elif col_type == "remark":
                row["remark"] = (
                    row["remark"] + " " + text
                ).strip()

            else:
                unknown_items.append(text)

        # ----------------------------------------------------
        # Fallback bộ phận
        # ----------------------------------------------------
        if not row["dept_src"]:
            candidates = []

            for item in line:
                text = normalize_ocr_text(item.get("text", ""))

                if not text:
                    continue

                if is_number_like(text):
                    continue

                if extract_stt(text) is not None:
                    continue

                candidates.append(text)

            # Ưu tiên text dài hơn vì thường là tên bộ phận.
            if candidates:
                row["dept_src"] = max(
                    candidates,
                    key=len,
                )

        # ----------------------------------------------------
        # Fallback remark cho text chưa phân loại.
        # Không lấy số vào remark.
        # ----------------------------------------------------
        if unknown_items:
            meaningful = [
                x for x in unknown_items
                if x and not is_number_like(x)
            ]

            if meaningful:
                existing = row["remark"]

                extra = " ".join(meaningful)

                row["remark"] = (
                    f"{existing} {extra}"
                ).strip()

        # ----------------------------------------------------
        # Không thêm dòng hoàn toàn rỗng.
        # ----------------------------------------------------
        has_data = any(
            str(row.get(k, "")).strip()
            for k in [
                "dept_src",
                "machines",
                "formal",
                "temp",
                "remark",
            ]
        )

        if has_data:
            rows.append(row)

    return rows


# ============================================================
# PARSE IMAGE
# ============================================================
def parse_attendance_image(image, mode):
    items = ocr_image(image)

    if not items:
        raise RuntimeError(
            "Không đọc được chữ từ ảnh này."
        )

    lines = group_ocr_lines(items)

    if not lines:
        raise RuntimeError(
            "OCR không tạo được dòng dữ liệu."
        )

    header_index = detect_header_line(lines)

    all_text = "\n".join(
        line_to_text(line)
        for line in lines
    )

    date_str = extract_date(all_text)

    if header_index is not None:
        title_lines = [
            line_to_text(line)
            for line in lines[:header_index]
        ]
    else:
        title_lines = []

    title_src = " ".join(
        x for x in title_lines if x
    ).strip()

    rows = parse_attendance_rows(
        lines,
        header_index,
    )

    title_tgt = (
        translate_text(
            title_src,
            mode,
            context="title",
        )
        if title_src
        else ""
    )

    for row in rows:
        if row.get("dept_src"):
            row["dept_tgt"] = translate_text(
                row["dept_src"],
                mode,
                context="department",
            )

        # Ghi chú cũng có thể chứa thuật ngữ nhà máy.
        if row.get("remark"):
            row["remark"] = translate_text(
                row["remark"],
                mode,
                context="remark",
            )

    return {
        "title_src": title_src,
        "title_tgt": title_tgt,
        "date_str": date_str,
        "rows": rows,
    }


# ============================================================
# PDF -> IMAGES
# ============================================================
def pdf_to_images(pdf_bytes):
    if fitz is None:
        raise RuntimeError(
            "Chưa cài PyMuPDF. Hãy cài pymupdf."
        )

    doc = fitz.open(
        stream=pdf_bytes,
        filetype="pdf",
    )

    images = []

    try:
        for page in doc:
            pix = page.get_pixmap(
                matrix=fitz.Matrix(2.5, 2.5),
                alpha=False,
            )

            image = Image.frombytes(
                "RGB",
                [pix.width, pix.height],
                pix.samples,
            )

            images.append(image)

    finally:
        doc.close()

    return images


# ============================================================
# MERGE DOCUMENTS
# ============================================================
def merge_parsed_documents(documents, mode):
    if not documents:
        return {
            "title_src": "",
            "title_tgt": "",
            "date_str": "",
            "rows": [],
        }

    merged = {
        "title_src": documents[0].get(
            "title_src",
            "",
        ),
        "title_tgt": documents[0].get(
            "title_tgt",
            "",
        ),
        "date_str": documents[0].get(
            "date_str",
            "",
        ),
        "rows": [],
    }

    next_stt = 1

    for doc in documents:
        for row in doc.get("rows", []):
            new_row = dict(row)

            new_row["stt"] = next_stt

            merged["rows"].append(new_row)

            next_stt += 1

    return merged


# ============================================================
# EXCEL
# ============================================================
def build_excel_from_json(data, mode):
    wb = Workbook()

    ws = wb.active
    ws.title = "Bảng chấm công"

    font_name = "Microsoft YaHei"

    orange_fill = PatternFill(
        fill_type="solid",
        fgColor="ED7D00",
    )

    thin_side = Side(
        style="thin",
        color="000000",
    )

    border = Border(
        left=thin_side,
        right=thin_side,
        top=thin_side,
        bottom=thin_side,
    )

    t_src = data.get("title_src", "")
    t_tgt = data.get("title_tgt", "")
    dt_str = data.get("date_str", "")
    rows = data.get("rows", [])

    title_parts = []

    if dt_str:
        title_parts.append(dt_str)

    if t_src:
        title_parts.append(t_src)

    full_title_src = " ".join(title_parts).strip()

    if t_tgt:
        full_title = (
            f"{full_title_src}\n{t_tgt}"
            if full_title_src
            else t_tgt
        )
    else:
        full_title = full_title_src

    ws.merge_cells("A1:F1")

    ws["A1"] = full_title

    ws["A1"].font = Font(
        name=font_name,
        size=13,
        bold=True,
    )

    ws["A1"].alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,
    )

    ws.row_dimensions[1].height = 48

    if mode == "Trung ➔ Việt":
        headers = [
            ("STT", "STT"),
            ("部门", "Bộ phận"),
            ("开几台机", "Số máy đang chạy"),
            ("正式工", "Công nhân chính thức"),
            ("临时工", "Công nhân thời vụ"),
            ("备注", "Ghi chú"),
        ]
    else:
        headers = [
            ("STT", "STT"),
            ("Bộ phận", "部门"),
            ("Số máy đang chạy", "开几台机"),
            ("Công nhân chính thức", "正式工"),
            ("Công nhân thời vụ", "临时工"),
            ("Ghi chú", "备注"),
        ]

    for col_idx, (top_h, bot_h) in enumerate(
        headers,
        start=1,
    ):
        cell = ws.cell(
            row=2,
            column=col_idx,
        )

        if top_h != bot_h:
            cell.value = f"{top_h}\n{bot_h}"
        else:
            cell.value = top_h

        cell.font = Font(
            name=font_name,
            size=10,
            bold=True,
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

        cell.fill = orange_fill
        cell.border = border

    ws.row_dimensions[2].height = 45

    current_row = 3

    for row in rows:
        values = [
            row.get("stt", ""),
            (
                f"{row.get('dept_src', '')}\n"
                f"{row.get('dept_tgt', '')}"
            ).strip(),
            row.get("machines", ""),
            row.get("formal", ""),
            row.get("temp", ""),
            row.get("remark", ""),
        ]

        for col_idx, value in enumerate(
            values,
            start=1,
        ):
            cell = ws.cell(
                row=current_row,
                column=col_idx,
                value=value,
            )

            cell.font = Font(
                name=font_name,
                size=10,
            )

            cell.alignment = Alignment(
                horizontal=(
                    "center"
                    if col_idx in [1, 3, 4, 5]
                    else "left"
                ),
                vertical="center",
                wrap_text=True,
            )

            cell.border = border

        ws.row_dimensions[current_row].height = 34

        current_row += 1

    widths = {
        "A": 8,
        "B": 32,
        "C": 16,
        "D": 22,
        "E": 20,
        "F": 28,
    }

    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    # Freeze header
    ws.freeze_panes = "A3"

    # Auto filter
    if current_row > 3:
        ws.auto_filter.ref = (
            f"A2:F{current_row - 1}"
        )

    output = io.BytesIO()

    wb.save(output)

    output.seek(0)

    return output


# ============================================================
# MEMORY EXPORT / IMPORT
# ============================================================
def build_memory_export():
    memory = load_translation_memory()

    return json.dumps(
        memory,
        ensure_ascii=False,
        indent=2,
    ).encode("utf-8")


def import_memory(uploaded_memory):
    try:
        data = json.loads(
            uploaded_memory.getvalue().decode(
                "utf-8"
            )
        )

        if not isinstance(data, dict):
            raise ValueError(
                "File memory không đúng định dạng."
            )

        data.setdefault("zh_vi", {})
        data.setdefault("vi_zh", {})

        save_translation_memory(data)

        return True

    except Exception as e:
        raise RuntimeError(
            f"Không thể nhập Translation Memory: {e}"
        )


# ============================================================
# SIDEBAR
# ============================================================
st.sidebar.header("⚙️ Cấu hình")

mode = st.sidebar.selectbox(
    "Chọn chiều dịch",
    [
        "Trung ➔ Việt",
        "Việt ➔ Trung",
    ],
)

st.sidebar.divider()

st.sidebar.subheader(
    "📚 Translation Memory"
)

memory = load_translation_memory()

st.sidebar.write(
    f"Trung → Việt: **{len(memory.get('zh_vi', {}))}** mục"
)

st.sidebar.write(
    f"Việt → Trung: **{len(memory.get('vi_zh', {}))}** mục"
)

memory_upload = st.sidebar.file_uploader(
    "Nhập Translation Memory JSON",
    type=["json"],
    key="memory_import",
)

if memory_upload is not None:
    try:
        import_memory(memory_upload)
        st.sidebar.success(
            "Đã nhập Translation Memory."
        )
    except Exception as e:
        st.sidebar.error(str(e))

st.sidebar.download_button(
    "📤 Xuất Translation Memory",
    data=build_memory_export(),
    file_name="translation_memory.json",
    mime="application/json",
)

st.sidebar.divider()

st.sidebar.subheader(
    "📊 Thống kê dịch"
)

stats = st.session_state.translation_stats

st.sidebar.write(
    f"📚 Glossary: **{stats['glossary']}**"
)

st.sidebar.write(
    f"🧠 Memory: **{stats['memory']}**"
)

st.sidebar.write(
    f"🔤 Argos: **{stats['argos']}**"
)

st.sidebar.write(
    f"🔢 Không dịch: **{stats['unchanged']}**"
)

if st.sidebar.button(
    "♻️ Xóa thống kê phiên này"
):
    st.session_state.translation_stats = {
        "glossary": 0,
        "memory": 0,
        "argos": 0,
        "unchanged": 0,
    }
    st.rerun()


# ============================================================
# FILE UPLOAD
# ============================================================
uploaded_file = st.file_uploader(
    "Tải lên tệp bảng chấm công",
    type=[
        "png",
        "jpg",
        "jpeg",
        "pdf",
    ],
    help=(
        "Hỗ trợ PNG, JPG, JPEG và PDF. "
        "PDF sẽ được chuyển thành ảnh trước khi OCR."
    ),
)


# ============================================================
# MAIN PROCESS
# ============================================================
if uploaded_file is not None:
    try:
        # ----------------------------------------------------
        # MODEL
        # ----------------------------------------------------
        with st.spinner(
            "Đang kiểm tra / chuẩn bị model dịch local..."
        ):
            initialize_translation_models()

        # ----------------------------------------------------
        # INPUT -> IMAGES
        # ----------------------------------------------------
        images = []

        if uploaded_file.type == "application/pdf":
            with st.spinner(
                "Đang chuyển PDF thành ảnh..."
            ):
                images = pdf_to_images(
                    uploaded_file.getvalue()
                )
        else:
            image = Image.open(uploaded_file)
            image.load()
            images = [image.copy()]

        if not images:
            raise RuntimeError(
                "Không tìm thấy trang/ảnh để xử lý."
            )

        # ----------------------------------------------------
        # PROCESS EACH PAGE
        # ----------------------------------------------------
        parsed_docs = []

        progress = st.progress(
            0,
            text="Đang bắt đầu xử lý...",
        )

        total = len(images)

        for idx, img in enumerate(images, start=1):
            progress.progress(
                int((idx - 1) / total * 100),
                text=(
                    f"Đang OCR + phân tích trang "
                    f"{idx}/{total}..."
                ),
            )

            parsed = parse_attendance_image(
                img,
                mode,
            )

            parsed_docs.append(parsed)

        progress.progress(
            100,
            text="Hoàn tất OCR và dịch.",
        )

        # ----------------------------------------------------
        # MERGE
        # ----------------------------------------------------
        final_data = merge_parsed_documents(
            parsed_docs,
            mode,
        )

        st.success(
            "Đã xử lý thành công."
        )

        # ----------------------------------------------------
        # SUMMARY
        # ----------------------------------------------------
        col1, col2, col3 = st.columns(3)

        with col1:
            st.metric(
                "Số trang",
                len(images),
            )

        with col2:
            st.metric(
                "Số dòng dữ liệu",
                len(final_data.get("rows", [])),
            )

        with col3:
            st.metric(
                "Translation Memory",
                len(
                    load_translation_memory().get(
                        memory_key(mode),
                        {},
                    )
                ),
            )

        # ----------------------------------------------------
        # PREVIEW
        # ----------------------------------------------------
        st.subheader(
            "👁️ Xem trước dữ liệu"
        )

        st.json(final_data)

        # ----------------------------------------------------
        # TABLE PREVIEW
        # ----------------------------------------------------
        preview_rows = []

        for row in final_data.get("rows", []):
            preview_rows.append(
                {
                    "STT": row.get("stt", ""),
                    "Bộ phận gốc": row.get(
                        "dept_src",
                        "",
                    ),
                    "Bộ phận dịch": row.get(
                        "dept_tgt",
                        "",
                    ),
                    "Số máy": row.get(
                        "machines",
                        "",
                    ),
                    "Chính thức": row.get(
                        "formal",
                        "",
                    ),
                    "Thời vụ": row.get(
                        "temp",
                        "",
                    ),
                    "Ghi chú": row.get(
                        "remark",
                        "",
                    ),
                }
            )

        if preview_rows:
            st.subheader(
                "📋 Bảng dữ liệu đã nhận diện"
            )

            st.dataframe(
                preview_rows,
                use_container_width=True,
                hide_index=True,
            )

        # ----------------------------------------------------
        # EXCEL
        # ----------------------------------------------------
        excel_data = build_excel_from_json(
            final_data,
            mode,
        )

        st.download_button(
            label=(
                "📥 Tải xuống Excel Song Ngữ"
            ),
            data=excel_data,
            file_name=(
                "bang_cham_cong_xu_ly_dong.xlsx"
            ),
            mime=(
                "application/vnd.openxmlformats-"
                "officedocument.spreadsheetml.sheet"
            ),
        )

    except Exception as e:
        st.error(
            f"Đã xảy ra lỗi khi xử lý tệp: {e}"
        )

        with st.expander(
            "🔎 Chi tiết lỗi"
        ):
            st.exception(e)
